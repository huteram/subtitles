# -*- coding: cp1250 -*-
# verze 2018-03
# created by Milan Hutera
# working with excel files


import openpyxl
import types
from operator import itemgetter


class XLS:
    def sheetList(self, Soubor):
        try:
            excel_document = openpyxl.load_workbook(Soubor)
            return excel_document.sheetnames
        except:
            return []
        
    
    def __HeadXLS(self, Source,Row = 1):
        Index = 1
        Head = {}
        for i in range(1,Source.max_column+1):
            if Source.cell(row=Row,column=i).value!= None:
                if (Source.cell(row=Row,column=i).value in Head)==False:
                    Aux = str(Source.cell(row=Row,column=i).value)                    
                    Head[Aux]=Index
                else:
                    Head[str(Index)] = Index
            else:
                Head[str(Index)] = Index
            Index = Index + 1
        
        return Head

    def __DataXLS(self, Source, Head, Row = 2):
        Data = []
        for i in range(Row,Source.max_row + 1 ):
            Aux = {}
            Add = False
            for ii in Head:
                
                Aux[ii] = Source.cell(row=i,column=Head[ii]).value
                if Aux[ii]!=None:
                    Add = True
                else:
                    Aux[ii] = ''
            if Add:
                Data.append(Aux)
                
        return Data

    def __DataXLSstr(self, Source, Head, Data = [], Row = 2):
        for i in range(Row,Source.max_row + 1 ):
            Aux = {}
            Add = False
            for ii in Head:
                
                Aux[ii] = Source.cell(row=i,column=Head[ii]).value
                if Aux[ii]!=None:
                    Add = True
                else:
                    Aux[ii] = ''
                Aux[ii] = str(Aux[ii])
            if Add:
                Data.append(Aux)
                
        return Data
    def RefSheets(self, Soubor, Sheet, ColumnList):
        ColumnListUse = []
        excel = self.ReadXLS(Soubor)
        head = excel[Sheet]['Head']

        if ColumnList[0] in head:
            for Column in ColumnList:
                ColumnListUse.append(head[Column])
        else:
            for Column in ColumnList:                    
                if not(type(Column) == int):
                    ColumnListUse.append( ord(Column.upper())-64)
                else:
                    ColumnListUse.append(Column)
                
        location = {}
        excel_document = openpyxl.load_workbook(Soubor)
        sheet = excel_document[Sheet]
        for i in range(2,sheet.max_row + 1 ):
            key = ''
            for Column in ColumnListUse:
                if type(sheet.cell(row=i,column=Column).value)!=type(None):  
                    key += sheet.cell(row=i,column=Column).value
            if key != '':
                if not(key in location):
                    location[key] = Sheet + "!" + chr(64 + ColumnListUse[0])+str(i)
                else:
                    location[key] = location[key].split(":")[0] + ":" + chr(64 + ColumnListUse[0])+str(i)
        excel_document.close()
        return location
     
    
    def SetHyperlink(self, Soubor, Sheet, ColumnList, Links):
        ColumnListUse = []        
        for Column in ColumnList:                    
            if not(type(Column) == int):
                ColumnListUse.append( ord(Column.upper())-64)
            else:
                ColumnListUse.append(Column)

        tempHyp = openpyxl.worksheet.hyperlink.Hyperlink
        tempHyp.ref = ""
            
        excel_document = openpyxl.load_workbook(Soubor)
        sheet = excel_document[Sheet]
        for i in range(2,sheet.max_row + 1 ):            
            key = ''
            for Column in ColumnListUse:
                
                if type(sheet.cell(row=i,column=Column).value)!=type(None):
                    
                    key += sheet.cell(row=i,column=Column).value
                    
            if key in Links:
                hyperlink = tempHyp()
                hyperlink.location = Links[key]
                sheet.cell(row=i,column=ColumnListUse[0]).hyperlink = hyperlink
                
        
        excel_document.save(Soubor)
        excel_document.close()

    def SetRows(self, Soubor, Sheet, ColumnList, Color = "DDDDDD", Color2 = "FFFFFF"):
        ColumnListUse = []        
        for Column in ColumnList:                    
            if not(type(Column) == int):
                ColumnListUse.append( ord(Column.upper())-64)
            else:
                ColumnListUse.append(Column)
                
        excel_document = openpyxl.load_workbook(Soubor)
        sheet = excel_document[Sheet]
        old = ''
        for Column in ColumnListUse:
            if type(sheet.cell(row=2,column=Column).value)!=type(None):  
                old += sheet.cell(row=2,column=Column).value
                
        color = False
        fill = openpyxl.styles.PatternFill("solid", fgColor=Color)
        fill_2 = openpyxl.styles.PatternFill("solid", fgColor=Color2)
        for i in range(2,sheet.max_row + 1 ):
            current = ''
            for Column in ColumnListUse:
                if type(sheet.cell(row=i,column=Column).value)!=type(None):  
                    current += sheet.cell(row=i,column=Column).value

                
            if old != current:
                color = not(color)
                old = current
            if color:
                for c in range(1,sheet.max_column+1):
                    cell = sheet.cell(row=i,column=c)
                    cell.fill = fill
            else:
                for c in range(1,sheet.max_column+1):
                    cell = sheet.cell(row=i,column=c)
                    cell.fill = fill_2
        excel_document.save(Soubor)
        excel_document.close()

    def NewSheet(self, SheetName):
        result = {}
        result[SheetName] = {}
        result[SheetName]['Head'] = {}
        result[SheetName]['Data'] = []
        return result
                               
    
    def ReadXLS(self, Soubor, Sheets = [], Head = {}, HeadRow = 1, DataRow = 2):
        workbook = {}
        excel_document = openpyxl.load_workbook(Soubor)
        
        
        if Sheets == []:
            Sheets = excel_document.sheetnames
        for i in  Sheets:
            workbook[i] = {}
            workbook[i]['Head'] = {}
            workbook[i]['Data'] = []
            
##            sheet = excel_document.get_sheet_by_name(i)
            sheet = excel_document[i]
            if Head == {}:
                workbook[i]['Head'] = self.__HeadXLS(sheet, Row = HeadRow)
            else:
                workbook[i]['Head'] = Head
            workbook[i]['Data'] = self.__DataXLS(sheet, workbook[i]['Head'], Row = DataRow )
            
        excel_document.close()
        return workbook
    
    def ReadXLSstr(self, Soubor, Sheets = [], Head = {}, HeadRow = 1, DataRow = 2):
        workbook = {}
        excel_document = openpyxl.load_workbook(Soubor)
        if Sheets == []:
            Sheets = excel_document.sheetnames
        for i in  Sheets:
            workbook[i] = {}
            workbook[i]['Head'] = {}
            workbook[i]['Data'] = []
            
##            sheet = excel_document.get_sheet_by_name(i)
            sheet = excel_document[i]
            if Head == {}:
                workbook[i]['Head'] = self.__HeadXLS(sheet, Row = HeadRow)
            else:
                workbook[i]['Head'] = Head
            workbook[i]['Data'] = self.__DataXLSstr(sheet, workbook[i]['Head'], workbook[i]['Data'], Row = DataRow )
        excel_document.close()
        return workbook
    def CSVHead(self, Head):
        aux = list(Head.values())
        aux.sort()
        result = list(range(aux[-1]))
        for i in Head:
            result[Head[i] - 1] = i
        return result

    def createHead(self, Head):
        """from array creates a Json """
        count = 1
        result = {}
        for i in Head:
            result[i] = count
            count += 1
        return result
        
            
    def SaveXLS(self, FileName, Workbook, Sheets = []):
        
        wb = openpyxl.Workbook()
        DefaultSheet = wb.sheetnames[0]
        lastHead = 0
        if Sheets == []:
            Sheets = list(Workbook.keys())
        for i in Sheets:
            if Sheets.index(i) == 0 and wb.sheetnames.count(DefaultSheet)>0:
                MySheet = wb.get_sheet_by_name(DefaultSheet)
                MySheet.title = i
            else:
                wb.create_sheet(i)
                MySheet = wb.get_sheet_by_name(i)
            for ii in Workbook[i]['Head']:
                MySheet.cell(row=1,column=Workbook[i]['Head'][ii]).value = ii
            row = 2
            for iii in Workbook[i]['Data']:
                for ii in iii:#Workbook[i]['Head']:
                    if ii in Workbook[i]['Head']:
##                        print(i,ii,iii[ii])
                        try:
                            MySheet.cell(row=row,column=Workbook[i]['Head'][ii]).value = iii[ii]
                        except:
                            print("trouble with this item to write it to Excel")
                            print(i,ii,iii[ii])
                    else:
                        if lastHead == 0:
                            for headItem in Workbook[i]['Head']:
                                lastHead = max(lastHead,Workbook[i]['Head'][headItem])
                        lastHead += 1
                        Workbook[i]['Head'][ii] = lastHead
                        MySheet.cell(row=1,column=Workbook[i]['Head'][ii]).value = ii
                        try:
                            MySheet.cell(row=row,column=Workbook[i]['Head'][ii]).value = iii[ii]
                        except:
                            print("trouble with this item to write it to Excel")
                            print(i,ii,iii[ii])
                            
                        
                row += 1    
        wb.save(FileName)
        wb.close()

if __name__ == '__main__':
    result = []
    import MujXLS
    zz = MujXLS.XLS()
    soubor = 'c:\Projects\CWL\CANconvertor_v3\sourceDbcXLS\CWL_v3.xlsx'
    wb = zz.ReadXLS('c:\Projects\CWL\CANconvertor_v3\sourceDbcXLS\CWL_v3.xlsx')
    print(wb['Tables']['Data'][0])
